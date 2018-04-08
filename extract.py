# -*- coding: utf-8 -*- #
import re
from openpyxl import load_workbook, Workbook

def trans_sheetname(sheetname):
    search_obj = re.search("A\d+", sheetname)
    if search_obj:
        return search_obj.group(0)
    return sheetname


def trans_sheetnames(rules):
    result = {}
    for name in rules:
        search_obj = re.search("A\d+", name)
        if search_obj:
            result[search_obj.group(0).encode("utf-8")] = rules[name]
    return result


#can have bug
def get_sequence(cell, last_cell, sheet_name):
    #横向累加
    if not cell[:1]==last_cell[:1]:
        this_row = cell[:1]
        last_row = last_cell[:1]
        this_row = ord(this_row)
        last_row = ord(last_row)
        result = ""
        last_row += 1
        while last_row <= this_row:
            result += ',cval(wb["%s"]["%s%d"])' % (sheet_name, chr(last_row), int(cell[1:]))
            last_row += 1
        return result
    #纵向累加
    else:
        this_row = cell[1:]
        last_row = last_cell[1:]
        while not this_row.isdigit():
            this_row = cell[1:]
        while not last_row.isdigit():
            last_row = last_cell[1:]
        this_row = int(this_row)
        last_row = int(last_row)
        result = ""
        last_row += 1
        while last_row <= this_row:
            result += ',cval(wb["%s"]["%s%d"])' % (sheet_name, cell[0], last_row)
            last_row += 1
        return result


def calculate_formula(wb, current_sheetname, formula):
    cell_pattern = re.compile(r'(\'?A[\d+][^\s]+?!)?([A-Z]{1,2}[0-9]+)')
    formula = formula.lstrip("=")
    result_str = ""

    last_cell = 0
    match_obj = re.search(cell_pattern, formula)
    

    while match_obj:
        sheet_name = match_obj.group(1)
        cell = match_obj.group(2)
        if not sheet_name:
            sheet_name = current_sheetname
        else:
            sheet_name = sheet_name[:-1]
            sheet_name = re.sub("\'","",sheet_name)
            sheet_name = trans_sheetname(sheet_name)

        value = 'cval(wb["%s"]["%s"])' % (sheet_name, cell)

        start, end = match_obj.span()
        if start == 1 and formula[0] == ":":
            result_str += get_sequence(cell, last_cell, sheet_name)
        else:
            result_str += formula[:start] + value
        formula = formula[end:]

        last_cell = cell

        match_obj = re.search(cell_pattern, formula)
    
    result_str += formula

    result_str = re.sub("(\d+)%", lambda x: "0."+x.group(1), result_str)

    result_str = re.sub("([^><])=", lambda x: x.group(1) + "==", result_str)

    result_str = re.sub("\$([A-Z]+)\$(\d+)", lambda x: 'cval(wb["%s"]["%s"])' % (current_sheetname, x.group(1)+x.group(2)), result_str)

    return result_str.encode("utf-8")


def getall_formula(wb):
    rules = {}
    for sheet_name in wb.sheetnames:
        rules[sheet_name] = {}
        for col in wb[sheet_name].iter_cols():
            for cell in col:
                value = cell.value
                if value and isinstance(value, (str, unicode)) and value[0] == "=":
                    (rules[sheet_name].setdefault(cell.column, {}))[cell.row] = calculate_formula(wb, trans_sheetname(sheet_name), value)
    return rules


wb = load_workbook(u"申报表软件要求original.xlsx", data_only=False)
sheet_names = wb.sheetnames

rules = getall_formula(wb)
rules = trans_sheetnames(rules)


rules_file = open("rules.txt", "w+")
rules_str = str(rules)
rules_file.write(rules_str)
rules_file.close()

#print(rules_str)
